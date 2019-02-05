from django.db.models import Max
from datetime import datetime
import xlrd
import openpyxl
import re
import os
import timeit

from .models import *   # Distr, DistrFile, DistrCategory, ExceptCode, Code, Price, Stock
from catalog.models import Category, Vendor

class Parsing:

    # Курс Нац банка
    def get_currency(self, date):
        from urllib import request
        import xmltodict
        import ssl
        ssl._create_default_https_context = ssl._create_unverified_context

        dt = date.strftime('%d.%m.%Y')
        file = request.urlopen('https://nationalbank.kz/rss/get_rates.cfm?fdate=' + dt)
        data = file.read()
        file.close()

        data = xmltodict.parse(data)
        i1 = list(data.items())
        i2 = list(i1[0][1].items())
        # i3 = list(i2[6][1].items())
        for e in i2[6][1]:
            i3 = list(e.items())
            if i3[1][1] == 'USD':
                rate = (i3[2][1])
                break

        return float(rate)

    # Считывает Excel файл и возвращает список
    def exl_list(self, file):
        ext = os.path.splitext(file.path)[1]
        if ext == '.xls':
            e = xlrd.open_workbook(file.path).sheet_by_index(0)
            excel_data = list()
            for rowno in range(e.nrows):
                row_data = list()
                for colno in range(e.ncols):
                    cell = e.cell(rowno, colno)
                    row_data.append(cell.value)
                excel_data.append(row_data)

        # print(excel_data)
        if ext == '.xlsx':
            # !!! На закоментировать строки 62-63 в openpyxl\styles\colors.py
            e = openpyxl.load_workbook(file).active
            excel_data = list()

            for row in e.iter_rows():                
                row_data = list()
                for cell in row:
                    row_data.append(cell.value)
                excel_data.append(row_data)

        return excel_data

    # Обработчик прайсов
    def process(self, source):
        a = timeit.default_timer()
        print('Процес!!!')

        count = 0   # Количество не исключённых товаров
        count_new = 0   # Количество товаров, которых нет в базе
        date = source.date  # Дата прайса

        raw = self.exl_list(source.file)    # Cписок данных из прайса

        distr = distributor[source.distr_id]()  # Специфика данных дистрибьютора
        # Получение даты
        new_date = distr.get_date(raw)
        if new_date:
            date = new_date

        last_date = Price.objects.filter(distr_id=source.distr_id).aggregate(Max('date'))['date__max']#.astimezone()
        if last_date:
            last_date = last_date.astimezone()
        else:
            last_date = date

        drate = None
        # Получение валюты
        if source.distr_id == 13 or source.distr_id == 14:
            dcur = DistrCurrency.objects.filter(distr_id=source.distr_id, date=date.date())
            if dcur:
                drate = float(dcur.first().rate)
            else:
                nbrk_cur = self.get_currency(date)
                drate = float(round(nbrk_cur * 1.005, 0))
                discur = DistrCurrency(symbol=1, rate=drate, date=date.date(), distr_id=source.distr_id)
                discur.save()


        dc_list = list()    # Список с id кодов
        # Перебор прайса
        for r in raw:
            code = distr.get_code(r)    # Код товара в прайсе дистрибьютора


            if code:
                exc = ExceptCode.objects.filter(distr_id=source.distr_id, code=code)    # Код исключённый из системы
                # Не исключён!
                if not exc:
                    count += 1
                    dc = Code.objects.filter(distr_id=source.distr_id, code=code)   # Код есть в системе
                    # Есть код
                    if dc:
                        dc = dc.first()
                        dc_list.append(dc.id)

                        # Работа по цене
                        p = distr.get_price(r, drate)  # Цена из прайса
                        # print(p)

                        already = Price.objects.filter(distr_id=source.distr_id, code_id=dc.id, date=date)
                        # Исключаем повторы
                        if not already:
                            pre_date = Price.objects.filter(distr_id=source.distr_id, code_id=dc.id, date__lt=date)
                            next_date = Price.objects.filter(distr_id=source.distr_id, code_id=dc.id, date__gt=date)

                            # Есть прайс с предыдущей датой
                            if pre_date:
                                ptype = 0
                                prd = pre_date.first()

                                # Цена изменилась
                                if p[0] != prd.price:
                                    # Цена была актуальна? Подразумевается, что "1" ещё и последняя цена! - подразумевается, но не работает!!!!
                                    if prd.type == 1:
                                        ptype = 1    # Актуализация цены
                                        # Деактивируем цену
                                        prd.type = 0
                                        prd.save()

                                    # Где актуальная цена? Если не последняя, то нужно "зачистить" следующие если цены =
                                    else:
                                        # Смотрим следующий прайс
                                        if next_date:
                                            nxd = next_date.first()
                                            # Удаляем совпадающие цены
                                            if p[0] == nxd.price:
                                                # Активная цена (присвоить)
                                                if nxd.type == 1:
                                                    ptype = 1

                                                prdel = Price.objects.get(pk=nxd.pk)
                                                prdel.delete()
                                        # Удаляем глючную цену
                                        else:
                                           Price.objects.filter(type=1, code_id=dc.pk).update(
                                               type=0)  # Зачишаем активные
                                           ptype = 1  # Крайняя - значит активе

                                    # Создаём нову цену
                                    newpr = Price(type=ptype, price=p[0], retail=p[1], date=date, distr_id=source.distr_id, code_id=dc.pk, product_id=dc.product_id)
                                    newpr.save()
                            # -  Есть прайс с предыдущей датой

                            # Самый ранний прайс
                            else:
                                ptype = 0
                                next_date = Price.objects.filter(distr_id=source.distr_id, code_id=dc.id, date__gt=date)
                                # Смотрим следующий прайс
                                if next_date:
                                    nxd = next_date.first()
                                    # Удаляем совпадающие цены
                                    if p[0] == nxd.price:
                                        # Активная цена (присвоить)
                                        if nxd.type == 1:
                                            ptype = 1

                                        prdel = Price.objects.get(pk=nxd.pk)
                                        prdel.delete()

                                # Последняя цена - значит актуальна
                                else:
                                    ptype = 1

                                # Добавляем первые цены
                                frstp = Price(type=ptype, price=p[0], retail=p[1], date=date, distr_id=source.distr_id, code_id=dc.pk, product_id=dc.product_id)
                                frstp.save()
                            # - Самый ранний прайс

                        # Тут работа по остаткам
                        stc = distr.get_stock(r)

                        if stc:
                            already = Stock.objects.filter(distr_id=source.distr_id, code_id=dc.id, date=date)
                            # Исключаем повторы
                            if not already:
                                pre_date = Stock.objects.filter(distr_id=source.distr_id, code_id=dc.id, date__lt=date)
                                next_date = Stock.objects.filter(distr_id=source.distr_id, code_id=dc.id, date__gt=date)

                                # Есть остатки с предыдущей датой
                                if pre_date:
                                    ptype = 0
                                    prd = pre_date.first()

                                    # Объём изменился
                                    if stc[0] != prd.value:
                                        # Остатки актуальны? Подразумевается, что "1" ещё и крайние остатки!
                                        if prd.type == 1:
                                            ptype = 1  # Актуализация объёма
                                            # Деактивируем остатки
                                            prd.type = 0
                                            prd.save()

                                        # Где актуальный? Если не последняя, то нужно "зачистить" следующие если объёмы равны
                                        else:
                                            # Смотрим следующий прайс
                                            if next_date:
                                                nxd = next_date.first()
                                                # Удаляем совпадающие цены
                                                if stc[0] == nxd.value:
                                                    # Актуальный объём (присвоить)
                                                    if nxd.type == 1:
                                                        ptype = 1

                                                    stdel = Stock.objects.get(pk=nxd.pk)
                                                    stdel.delete()

                                        # Создаём новвый запас
                                        newstock = Stock(type=ptype, var=stc[1], value=stc[0], date=date,
                                                      distr_id=source.distr_id, code_id=dc.pk, product_id=dc.product_id)
                                        newstock.save()
                                # -  Есть прайс с предыдущей датой

                                # Самый ранний прайс
                                else:
                                    ptype = 0
                                    next_date = Stock.objects.filter(distr_id=source.distr_id, code_id=dc.id,
                                                                     date__gt=date)
                                    # Смотрим следующий прайс
                                    if next_date:
                                        nxd = next_date.first()
                                        # Удаляем совпадающие объёмы
                                        if stc[0] == nxd.value:
                                            # Активная цена (присвоить)
                                            if nxd.type == 1:
                                                ptype = 1

                                            stockdel = Stock.objects.get(pk=nxd.pk)
                                            stockdel.delete()

                                    # Последний объём - значит актуальный
                                    else:
                                        ptype = 1

                                    # Добавляем первые объёмы
                                    newstock = Stock(type=ptype, var=stc[1], value=stc[0], date=date,
                                                     distr_id=source.distr_id, code_id=dc.pk, product_id=dc.product_id)
                                    newstock.save()
                                # - Самый ранний прайс

                    # Кодв нет в базе
                    else:
                        count_new += 1

        # Обновить наличие товара
        if date >= last_date:
            Code.objects.filter(type=1, distr_id=source.distr_id).update(type=0)    # Деактивировать всё
            Code.objects.filter(id__in=dc_list).update(type=1)  # Наличие товара по коду из последнего прайса
            print("Наличие обновлено!!!")


        # Обновить статус прайса
        file_type = 1
        if count_new:
            file_type = 3
        else:
            print('Прайс обработан!')

        source.type = file_type
        source.date = date
        source.count = count
        source.count_new = count_new
        source.save()



        print(date, count, count_new)
        print(timeit.default_timer() - a)
        return count_new

    # Товары не в базе
    def get_uncode(self, source, step=False):
        a = timeit.default_timer()
        file = source.file
        # Подменяем файл у Компортала
        if source.distr_id == 17:
            fg = DistrFile.objects.filter(type=1, distr_id=17, date=source.date)
            if fg:
                file = fg.first()
                file = file.file

        raw = self.exl_list(file)  # Cписок данных из прайса
        distr = distributor[source.distr_id]()  # Специфика данных дистрибьютора

        dc_q = Code.objects.values_list('code', flat=True).filter(distr_id=source.distr_id)
        dc_list = set(dc_q)    # Список кодов дистрибьютора
        dce_q = ExceptCode.objects.values_list('code', flat=True).filter(distr_id=source.distr_id)
        dce_list = set(dce_q)  # Список исключённых кодов дистрибьютора

        category = Category.objects.get(pk=10)
        cc = ''     # Оригинальное наименование категории
        code_list = list()
        ok = True
        for r in raw:
            code = distr.get_code(r)  # Код товара в прайсе дистрибьютора
            if code:
                #exc = ExceptCode.objects.filter(distr_id=source.distr_id, code=code)  # Код исключённый из системы
                # Не исключён!
                if code not in dce_list:
                    #dc = Code.objects.filter(distr_id=source.distr_id, code=code)  # Код есть в системе
                    # Нет кода в базе
                    if code not in dc_list:
                        name = distr.get_name(r)  # Текстовые данные товара
                        cat = distr.get_cat(r)  # Для дистров у которых категория в товаре
                        if cat:
                            cc = cat
                            dcat = DistrCategory.objects.filter(distr_id=source.distr_id, name=cat)
                            if dcat:
                                catc = dcat.first()
                                category = catc.category
                                # Исключаем товары, которые входят в заблокированную категорию
                                if catc.type == 2:
                                    exc = ExceptCode(type=1, code=code, description=name[2], comment=name[3],
                                                     distr_id=source.distr_id)
                                    exc.save()
                                    # print(name[2])
                            else:
                                category = Category.objects.get(pk=10)
                                # Вписываем новые категории
                                c = DistrCategory.objects.filter(distr_id=source.distr_id, name=cat)
                                if not c:
                                    ins = DistrCategory(type=3, name=cat, category_id=10, distr_id=source.distr_id)
                                    ins.save()


                        raw_dict = {
                            'code': code,
                            'category': category,
                            'vendor': distr.get_vendor(r),
                            'iarticle': name[0],
                            'article': name[1],
                            'name': name[2],
                            'description': name[3],
                            'cat': cc,
                        }
                        code_list.append(raw_dict)  # Собираем всё в кучу
                        ok = False
                        if step:
                            break

            # Данные между товарами
            else:
                cat = distr.get_cat(r)  # Для дистров у которых категории между товарами
                if cat:
                    cc = cat
                    dcat = DistrCategory.objects.filter(type=1, distr_id=source.distr_id, name=cat)
                    if dcat:
                        catc = dcat.first()
                        category = catc.category
                    else:
                        category = Category.objects.get(pk=10)
                        # Вписываем новые категории
                        c = DistrCategory.objects.filter(distr_id=source.distr_id, name=cat)
                        if not c:
                            ins = DistrCategory(type=3, name=cat, category_id=10, distr_id=source.distr_id)
                            ins.save()
        if ok:
            print("Запуск!!!")
            self.process(source)

        print(timeit.default_timer() - a)
        return code_list

    def go(self):
        flist = DistrFile.objects.filter(type=2)
        for f in flist:
            self.process(f)
        return len(flist)

# ===========================================================+

class Victory:
    def __init__(self):
        print('Я Виктори!')

    def get_date(self, list):
        i = 2
        if not list[0][i]:
            i = 1
        day = re.findall(r'\d{2}.\d{2}.\d{4}', list[0][i])[0]
        time = re.findall(r'\d{2}:\d{2}', list[0][i])[0]
        return datetime.strptime(day + time, '%d.%m.%Y%H:%M').astimezone()

    def get_code(self, line):
        code = None
        if line[0]:
            code = str(line[0])
        if code == 'ID:     ':
            code = 'ID:'

        return code

    def get_price(self, line, cur=False):
        price = (float(line[3]), None)
        return price

    def get_stock(self, list):
        return None

    def get_cat(self, list):
        cat = None
        if not list[0]:
            cat = list[1]
        return cat

    def get_name(self, line):
        article = ''
        if line[1]:
            article = line[1].replace(' oem', '')
        name = ('', article, line[2], line[2])
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            find = re.search(r'\b{}\b'.format(vend.vendor), line[2], re.IGNORECASE)
            if find:
                vnd = Vendor.objects.get(pk=vend.pk)
                vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

class Redington:
    def __init__(self):
        print('Я Redington!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[4]:
            code = str(line[1])
            count = len(code)
            nols = '0'
            for i in range(count, 12):
                nols += '0'
            code = nols+code.replace('.0', '')

        return code

    def get_price(self, line, cur=False):
        p = 0
        if line[7]:
            p = float(line[7])
        price = (p, None)
        return price

    def get_stock(self, line):
        stock = None
        if line[6]:
            stock = str(line[6])
            if stock[0] == '<':
                var = 1
                val = stock.replace('< ', '')
            elif stock[0] == '>':
                var = 2
                val = stock.replace('> ', '')
            else:
                var = 0
                val = stock.replace('.0', '')
            stock = (int(val), var)

        return stock

    def get_cat(self, line):
        cat = None
        if not line[4]:
            cat = line[1]
        return cat

    def get_name(self, line):
        if type(line[2]) is float:
            line[2] = str(line[2]).replace('.0', '')
        name = (str(line[2]), str(line[3]), str(line[4]), str(line[4]))
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            find = re.search(r'\b{}\b'.format(vend.vendor), line[4], re.IGNORECASE)
            if find:
                vnd = Vendor.objects.get(pk=vend.pk)
                vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

class Arenas:
    def __init__(self):
        print('Я Arena S!')

    def get_date(self, line):
        day = re.findall(r'\d{2}.\d{2}.\d{4}', line[2][1])[0]
        time = re.findall(r'\d{2}:\d{2}', line[2][1])[0]
        return datetime.strptime(day + time, '%d.%m.%Y%H:%M').astimezone()

    def get_code(self, line):
        code = None
        if line[7]:
           code = str(line[1])

        return code

    def get_price(self, line, cur=False):
        p = 0
        r = 0
        if line[6]:
            p = line[6]
        if line[7] and type(line[7]) is not str:
            r = line[7]


        print(p, type(p), r, type(r))
        price = (float(p), float(r))
        return price

    def get_stock(self, line):
        stock = None
        return stock

    def get_cat(self, line):
        cat = None
        if line[4]:
            cat = line[3]
        return cat

    def get_name(self, line):
        name = ('', '', str(line[5]), str(line[9]))
        return name

    def get_vendor(self, line):
        vend_list = list()
        if line[4] == "HEWLETT PACKARD":
            line[4] = 'HP'
        if line[4] == "REMAX (China)":
            line[4] = 'Remax'
        if line[4] == "WESTERN DIGITAL":
            line[4] = 'WD'


        vend = Vendor.objects.filter(type=1, vendor__icontains=line[4])
        if vend:
            vend = vend.first()
            vnd = Vendor.objects.get(pk=vend.pk)
            vend_list.append(vnd)
        else:
            vnd = Vendor.objects.get(pk=10)
            vend_list.append(vnd)

        return vend_list

class Moon:
    def __init__(self):
        print('Я Мун!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[3]:
            code = str(line[1])
            count = len(code)
            nols = '0'
            for i in range(count, 12):
                nols += '0'
            code = nols + code.replace('.0', '')

        return code

    def get_price(self, line, cur=False):
        price = (float(line[3]), float(line[4]))
        return price

    def get_stock(self, list):
        return None

    def get_cat(self, list):
        cat = None
        if not list[3]:
            day = re.findall(r'Цены указаны на', list[2])
            if not day:
                cat = list[2]
        return cat

    def get_name(self, line):
        name = ('', '', line[2], line[2])
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            find = re.search(r'\b{}\b'.format(vend.vendor), line[2], re.IGNORECASE)
            if find:
                vnd = Vendor.objects.get(pk=vend.pk)
                vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list



class Comportal:
    def __init__(self):
        print('Я Ком Портал!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[1]:
            code = str(line[0]).replace('.0', '')

        print(line[0])
        return code

    def get_price(self, line, cur=False):
        price = (float(line[4]), None)
        return price

    def get_stock(self, line):
        val = str(line[5]).replace('.0', '')
        stock = (int(val), 0)
        return stock

    def get_cat(self, list):
        cat = None
        if not list[2]:
            cat = str(list[1]).replace(' ', '')
        return cat

    def get_name(self, line):
        name = (line[2], '', line[3], line[3])
        #print(name)
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            if line[3]:
                find = re.search(r'\b{}\b'.format(vend.vendor), line[3], re.IGNORECASE)
                if find:
                    vnd = Vendor.objects.get(pk=vend.pk)
                    vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

# Стврый парсер
class Comportal_old:
    def __init__(self):
        print('Я Ком Портал!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[2]:
            code = str(line[1]).replace('.0', '')

        return code

    def get_price(self, line, cur=False):
        price = (float(line[4]), None)
        return price

    def get_stock(self, line):
        val = str(line[5]).replace('.0', '')
        stock = (int(val), 0)
        return stock

    def get_cat(self, list):
        cat = None
        if not list[2]:
            cat = str(list[1]).replace(' ', '')
        return cat

    def get_name(self, line):
        name = (line[2], '', line[3], line[3])
        #print(name)
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            if line[3]:
                find = re.search(r'\b{}\b'.format(vend.vendor), line[3], re.IGNORECASE)
                if find:
                    vnd = Vendor.objects.get(pk=vend.pk)
                    vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

class Marvel:
    def __init__(self):
        print('Я Марвел!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[5]:
            code = line[5]

        return code

    def get_price(self, line, cur):
        p = 0
        if type(line[9]) is not str:
            p = round(line[9], 2)
            if line[10] == 'USD':
                p = round(p * cur, 2)

        price = (float(p), None)
        return price

    def get_stock(self, line):
        val = 0
        var = 0
        if line[12]:
            stc = str(line[12])
            if stc[-1:] == '+':
                val = stc.replace('+', '')
                var = 1
            else:
                val = stc.replace('.0', '')

        stock = (int(val), var)
        return stock

    def get_cat(self, line):
        cat = None
        if line[2]:
            cat = line[2]
        return cat

    def get_name(self, line):
        name = (str(line[5]), '', str(line[6]), str(line[6]))
        return name

    def get_vendor(self, line):
        vend_list = list()
        vend = Vendor.objects.filter(type=1, vendor__icontains=line[4])
        if vend:
            vend = vend.first()
            vnd = Vendor.objects.get(pk=vend.pk)
            vend_list.append(vnd)
        else:
            vnd = Vendor.objects.get(pk=10)
            vend_list.append(vnd)

        return vend_list

class Vstrade:
    def __init__(self):
        print('Я VS Trade!')

    def get_date(self, list):
        day = re.findall(r'\d{2}.\d{2}.\d{4}', list[0][0])[0]
        return datetime.strptime(day + '00:00', '%d.%m.%Y%H:%M').astimezone()

    def get_code(self, line):
        code = None
        if line[4]:
            code = str(line[0])
        # print(code)
        return code

    def get_price(self, line, cur):
        p = 0
        r = 0
        if line[4]:
            p = round(line[4], 2)
            p = round(p * cur, 2)
        if line[5]:
            r = round(line[5], 2)
            r = round(r * cur, 2)

        price = (float(p), float(r))
        return price

    def get_stock(self, line):
        val = 0
        var = 0
        if line[6]:
            val = line[6]

        stock = (int(val), var)
        return stock

    def get_cat(self, list):
        cat = None
        if type(list[0]) is str:
            if list[0][0:11] != 'Прайс-лист:':
                cat = list[0]

        return cat

    def get_name(self, line):
        name = (str(line[1]), '', str(line[2]), str(line[2]))
        return name

    def get_vendor(self, line):
        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            find = re.search(r'\b{}\b'.format(vend.vendor), line[2], re.IGNORECASE)
            if find:
                vnd = Vendor.objects.get(pk=vend.pk)
                vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

class Alstyle:
    def __init__(self):
        print('Я Al-style!')

    def get_date(self, list):
        return None

    def get_code(self, line):
        code = None
        if line[2]:
            code = str(line[0])

        return code

    def get_price(self, line, cur=False):
        price = (float(line[2]), None)
        return price

    def get_stock(self, line):
        stock = None
        if line[4]:
            stock = str(line[4])
            if stock[0] == '<':
                var = 1
                val = stock.replace('<', '')
            elif stock[0] == '>':
                var = 2
                val = stock.replace('>', '')
            else:
                var = 0
                val = stock.replace('.0', '')
            stock = (int(val), var)

        return stock

    def get_cat(self, list):
        cat = None
        if type(list[0]) is str:
            cat = list[0]

        return cat

    def get_name(self, line):
        name = ('', '', str(line[1]), str(line[1]))
        return name

    def get_vendor(self, line):


        vend_list = list()
        v = Vendor.objects.filter(type=1)
        for vend in v:
            if vend.vendor == 'WD':
                vend.vendor = 'Western Digital'
            find = re.search(r'\b{}\b'.format(vend.vendor), line[1], re.IGNORECASE)
            if find:
                vnd = Vendor.objects.get(pk=vend.pk)
                vend_list.append(vnd)

        if not vend_list:
            vend_list.append(Vendor.objects.get(pk=10))
        return vend_list

distributor = {
    11: Victory,
    12: Redington,
    13: Vstrade,
    14: Marvel,
    15: Arenas,
    16: Moon,
    17: Comportal,
    18: Alstyle,
}










